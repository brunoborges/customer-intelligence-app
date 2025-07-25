#!/usr/bin/env python3
"""
Script to generate customer profiles using OpenAI GPT-4.
Reads customer data from nudge_customers.xlsx and profile data from profile_book.pdf,
then generates detailed profiles for each customer.
"""

import os
import sys
import time
import re
from typing import List, Dict, Optional
import PyPDF2
import openpyxl
from openai import OpenAI
import os

def load_openai_key() -> str:
    """Load OpenAI API key from environment variable."""
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        print("Error: OPENAI_API_KEY environment variable is not set!")
        sys.exit(1)
    return api_key

def extract_pdf_text(pdf_path: str) -> str:
    """Extract text from PDF file."""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ''
            for page in pdf_reader.pages:
                text += page.extract_text() + '\n'
        return text
    except Exception as e:
        print(f"Error reading PDF: {e}")
        sys.exit(1)

def find_person_in_profile_book(first_name: str, last_name: str, profile_text: str) -> str:
    """
    Search for a person's profile in the profile book text.
    Returns the relevant section of text for that person.
    """
    # Try different variations of the name
    name_variations = [
        f"{first_name} {last_name}",
        f"{first_name.upper()} {last_name.upper()}",
        f"{last_name}, {first_name}",
        f"{last_name.upper()}, {first_name.upper()}",
        first_name,
        last_name
    ]
    
    best_match = ""
    best_score = 0
    
    for name_var in name_variations:
        if name_var.lower() in profile_text.lower():
            # Find the context around the name
            pattern = re.compile(rf"(.{{0,500}}){re.escape(name_var)}(.{{0,1000}})", re.IGNORECASE | re.DOTALL)
            matches = pattern.findall(profile_text)
            
            for before, after in matches:
                context = before + name_var + after
                # Score based on length and relevance indicators
                score = len(context)
                if any(keyword in context.lower() for keyword in ['education', 'experience', 'work', 'position', 'company']):
                    score += 200
                
                if score > best_score:
                    best_score = score
                    best_match = context
    
    return best_match.strip() if best_match else f"No specific profile found for {first_name} {last_name} in the profile book."

def generate_customer_profile(client: OpenAI, first_name: str, last_name: str, city: str, profile_book_data: str) -> str:
    """Generate a customer profile using OpenAI."""
    
    prompt = f"""
You are a customer analytics expert. Based on the following information, create a detailed customer profile for {first_name} {last_name} who lives in {city}.

Profile Book Data (if available):
{profile_book_data}

Generate a comprehensive customer profile that includes:

1. **Demographics & Location**: Based on living in {city}
2. **Purchasing Habits**: Create realistic shopping patterns and preferences
3. **Interests & Lifestyle**: Infer interests based on available information and city demographics
4. **Financial Profile**: Generate a realistic credit score (300-850) and spending capacity
5. **Brand Preferences**: Suggest likely brand affinities
6. **Shopping Behavior**: Online vs in-store preferences, seasonal patterns
7. **Communication Preferences**: Preferred channels and messaging style

Make the profile realistic and detailed (2-3 paragraphs), incorporating any professional background or education information from the profile book data if available. If no specific data is available for this person, create a believable profile based on demographic patterns for someone in {city}.

Include a credit score at the end in this format: "Credit Score: XXX"

Respond with just the profile text, no additional formatting or headers.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are an expert customer analytics specialist who creates detailed, realistic customer profiles."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=500,
            temperature=0.7
        )
        
        return response.choices[0].message.content.strip()
        
    except Exception as e:
        print(f"Error generating profile for {first_name} {last_name}: {e}")
        return f"Error generating profile: {str(e)}"

def load_customers_from_excel(file_path: str) -> List[Dict]:
    """Load customer data from Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        customers = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_num in range(2, sheet.max_row + 1):
            row_data = {}
            for col_num, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                row_data[header] = cell_value
            customers.append(row_data)
        
        return customers
        
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)

def update_excel_with_profiles(file_path: str, customers: List[Dict]):
    """Update Excel file with generated profiles."""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Find the profile column index
        headers = [cell.value for cell in sheet[1]]
        profile_col_index = headers.index('profile') + 1
        
        # Update each row with the profile
        for row_num, customer in enumerate(customers, 2):
            if customer.get('profile'):
                sheet.cell(row=row_num, column=profile_col_index, value=customer['profile'])
        
        # Save the file
        wb.save(file_path)
        print(f"Successfully updated {file_path} with generated profiles!")
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def main():
    """Main function to process all customers and generate profiles."""
    print("Starting customer profile generation...")
    
    # Initialize OpenAI client
    api_key = load_openai_key()
    client = OpenAI(api_key=api_key)
    
    # Load profile book data
    print("Loading profile book data...")
    profile_book_text = extract_pdf_text('profile_book.pdf')
    
    # Load customers from Excel
    print("Loading customer data...")
    customers = load_customers_from_excel('nudge_customers.xlsx')
    print(f"Found {len(customers)} customers to process")
    
    # Generate profiles for each customer
    for i, customer in enumerate(customers, 1):
        first_name = customer.get('first_name', '')
        last_name = customer.get('last_name', '')
        city = customer.get('city', '')
        
        print(f"Processing {i}/{len(customers)}: {first_name} {last_name} from {city}")
        
        # Skip if profile already exists
        if customer.get('profile'):
            print(f"  Profile already exists, skipping...")
            continue
        
        # Find person's data in profile book
        person_profile_data = find_person_in_profile_book(first_name, last_name, profile_book_text)
        
        # Generate profile using OpenAI
        profile = generate_customer_profile(client, first_name, last_name, city, person_profile_data)
        customer['profile'] = profile
        
        print(f"  Generated profile ({len(profile)} characters)")
        
        # Rate limiting - wait 1 second between API calls
        time.sleep(1)
    
    # Update Excel file with generated profiles
    print("Updating Excel file...")
    update_excel_with_profiles('nudge_customers.xlsx', customers)
    
    print("Profile generation completed successfully!")

if __name__ == "__main__":
    main()